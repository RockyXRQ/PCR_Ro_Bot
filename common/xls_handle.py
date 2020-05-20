from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Color, Font


class xlsHandle:
    wb = load_workbook(filename='影都第二次公会战数据.xlsx')


def wb_update():
    xlsHandle.wb.save('影都第二次公会战数据.xlsx')
    xlsHandle.wb = load_workbook(filename='影都第二次公会战数据.xlsx')


def xls_create_user(user_nickname: str):
    ws = xlsHandle.wb.copy_worksheet(xlsHandle.wb['Template'])
    ws.title = user_nickname
    ws['B1'] = user_nickname
    wb_update()


def xls_damage_append(user_nickname: str, team_info: str, damage: int, day: int, isKill: bool):
    ft = Font(color=colors.RED)

    ws = xlsHandle.wb[user_nickname]
    if not ws.cell(4, 2+day):
        ws.cell(4, 2+day).value = team_info
        ws.cell(5, 2+day).value = damage
        if isKill:
            ws.cell(5, 2+day).font = ft
    elif not ws.cell(6, 2+day):
        ws.cell(6, 2+day).value = team_info
        ws.cell(7, 2+day).value = damage
        if isKill:
            ws.cell(7, 2+day).font = ft
    else:
        ws.cell(8, 2+day).value = team_info
        ws.cell(9, 2+day).value = damage
        if isKill:
            ws.cell(9, 2+day).font = ft
    wb_update()


def xls_on_tree(user_nickname: str):
    ws = xlsHandle.wb[user_nickname]
    ws.cell(5, 12).value += 1
    wb_update()


def xls_save_tree(user_nickname: str):
    ws = xlsHandle.wb[user_nickname]
    ws.cell(6, 12).value += 1
    wb_update()


def xls_get_sheets():
    return xlsHandle.wb.sheetnames


def xls_get_boss_list():
    ws = xlsHandle.wb['Boss']
    return [[ws['A1'], ws['B1']], [ws['A2'], ws['B2']], [
        ws['A3'], ws['B3']], [ws['A4'], ws['B4']], [ws['A5'], ws['B5']]]


def xls_update_boss_list(boss: int, damage: int):
    ws = xlsHandle.wb['Boss']
    ws['B'+boss] = max(0, ws['B'+boss]-damage)


def xls_restore_boss_list():
    ws = xlsHandle.wb['Boss']
    ws['B1'] = 6000000
    ws['B2'] = 8000000
    ws['B3'] = 10000000
    ws['B4'] = 12000000
    ws['B5'] = 20000000
